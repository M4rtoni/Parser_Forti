#!/bin/python
# -*- coding: utf-8 -*-

"""################################################################################
#
# File Name         : parser_forti.py
# Created By        : Florian MAUFRAIS
# Contact           : florian.maufrais@nxosecurity.com
# Creation Date     : april  13th, 2015
# Version           : 1.2.4
# Last Change       : June 20th, 2017 at 10:42
# Last Changed By   : Florian MAUFRAIS
# Purpose           : This programm could parse backup files of a
#                     Fortigate appliance to a JSON/dict and XLSX structure
#
#
################################################################################"""

__version__ = (1,2,4)
__all__ = ['parse','Parser','Parsed_to_xls','prepare','build_xls']

################################################################################

import argparse, sys, os

LOCAL_DIR = os.path.dirname(os.path.abspath(__file__))
#LOCAL_DIR = os.getcwd()

class Check_path(argparse.Action):
    def __init__(self, option_strings, dest, nargs=None, type=str, **kwargs):
        if nargs is not None:
            raise ValueError("nargs not allowed")
        if type is not str:
            raise ValueError("type must be an str")
        super(Check_path, self).__init__(option_strings, dest, **kwargs)
    def __call__(self, parser, namespace, values, option_string=None):
        if os.path.isdir(values):
            setattr(namespace, self.dest, os.path.dirname(values)+os.sep)

parser = argparse.ArgumentParser(description='Process some Fortigate configuration files.')
parser.add_argument('--dir', dest='dir', action=Check_path,
    default='./', help='searching folder (default: ./)')
parser.add_argument('--files', dest='files', action='store', nargs='+',
    default='*.conf', help='type off file search (default: *.conf)')
parser.add_argument('-json', dest='JSON', action='store_true',
    default=False, help='write result in a JSON file (default: False)')
parser.add_argument('-xlsx', dest='XLSX', action='store_true',
    default=False, required='-webfilter' in sys.argv[1:] or 
    '--keys' in sys.argv[1:],
    help='write result in a XLSX file (default: False)')
parser.add_argument('-webfilter', dest='webfilter', action='store_true',
    default=False, required='--profile' in sys.argv[1:],
    help='remplace webfilter category in xlsx file (require xlsx, search '+\
    '"webfilter_profiles.json", default: False)')
parser.add_argument('--profile', dest='profile', action='store',
    default=LOCAL_DIR+os.sep+'webfilter_profiles.json', 
    help='change file for webfilter category (default: webfilter_profiles.json)')
parser.add_argument('--keys',dest='keys', action='store', nargs='+',
    default=['firewall','vpn','proxy','webfilter','router','ips','system',
    'application','antivirus'], help='Change default keys use to select '+\
    'main config descriptor (default list : firewall, vpn, proxy, webfilter, '+\
    'router, ips, system, application, antivirus)')
if __name__ == '__main__':
    args = parser.parse_args(sys.argv[1:])

import shlex, json

# Openpyxl (c) 2017
# Author : Eric Gazoni, Charlie Clark
# Licence : MIT/Expat
# Version 2.4.7

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

class Parser(dict):
    def __init__(self,config):
        self.headers = []
        if type(config) is str:
            for conf in config.split('\n'):
                if conf.startswith('#'):
                    self.headers.append(conf)
                else: 
                    break
            index = sum([len(header) for header in self.headers])+len(self.headers)
            self.__config = parse(config[index:], res = {})
            self.__config.update({'headers':self.headers})
        elif type(config) is dict:
            self.__config = config
    def __repr__(self):
        return self.__config.__repr__()
    def __getitem__(self,*args):
        return self.__config.__getitem__(*args)
    def keys(self):
        return self.__config.keys()
    def prepare(self, key = None):
        if key == None:
            self.__prepared = prepare(self.__config)
        else:
            self.__prepared = prepare(self.__config.__get__(key,None))
        return self.__prepared
    def build_xls(self, key = None):
        if key == None:
            self.__built_xls = build_xls(self.__prepared)
        else:
            self.__built_xls = build_xls(self.__prepared.__get__(key,None))
    def Parsed_to_xls(self, key, name, save = False):
        self.__workbook = Parsed_to_xls(self.__prepared, key, name, save)
        return self.__workbook

class Webfilter_category:
    def __init__(self, id, name, group, version=None):
        self.id = id
        self.group = group
        self.name = name
        self.version = version
    def __repr__(self):
        return str('id = %r, name = %r, group = %r' % (self.id, self.name, self.group) + ('',', version = %r' % self.version)[self.version != None])

def parse(config,end=('end','next'),res={},titles=('set','unset')):
    """ Function that take Fortigate backup configuration file and return
    a JSON/dict structure where every key is a JSON/dict elements 
    with an higer level
    
    Here is an example:
        >>> config = 'config system interface\\n'+\\
        ... '    edit "port 1"\\n'+\\
        ... '        set vdom "root"\\n'+\\
        ... '        set ip 192.168.1.254 255.255.255.0\\n'+\\
        ... '        set allowaccess ping\\n'+\\
        ... '        set type physical\\n'+\
        ... '        set alias "Internet access"\\n'+\\
        ... '        set snmp-index 1\\n'+\\
        ... '    next\\n'+\\
        ... 'end'
        >>> result = parse(config)
        >>> print result.keys()
        ['config system interface']
        >>> print result['config system interface']['edit "port 1"']['ip']
        {'set': ['192.168.1.254', '255.255.255.0']}
        >>> print result['config system interface']['edit "port 1"']['type']
        {'set': ['physical']}
    
    return a dict
    """
    
    #
    # On serape chaque entrée, en prenant compte des entrées multi-lignes
    #
    
    _config = config.split('\n')
    __config = []
    _c = ''
    for c in _config:
        _c += c
        __c = True
        try:
            ___c = shlex.split(_c)
        except:
            __c = False
            _c += '\n'
        if __c:
            __config.append(_c)
            _c = ''
    
    _config = __config
    index = 0
    _old_index = 0
    
    #
    # On boucle a chaque bloc
    #
    
    while index < len(_config):
        
        #
        # On prépare le titre du bloc
        #
        
        title = _config[index]
        _end = None
        
        #
        # On prend en compte l'indentation
        #
        
        if title.startswith(' '*4):
            _end = end
        
        while title.startswith(' '*4):
            end = tuple((' '*4+e for e in end))
            title = title[4:]
        
        index+=1
        
        #
        # On cherche la fin du bloc
        #
        
        if True in [_config[index:].__contains__(e) for e in end]:
            stop = max([_config.index(e,index) for e in end if _config[index:].__contains__(e)])
        else:
            stop = len(_config)
        
        #
        # Identification du bloc à travailler
        #
        
        working = _config[index:stop]
        
        #
        # On retire l'indentation
        #
        
        working = [w.replace(' '*4,'',1) for w in working] 
        while set([w[:4] for w in working if not w in end]) == set([' '*4]):
            working = [w.replace(' '*4,'',1) for w in working]
        
        rep = {}
        rem = []
        
        #
        # On cherche les titres de variable (set, unset, edit)
        # "edit" est un titre ambigu, il peut servir pour un titre de variable ou de bloc
        #
        
        for i in range(len(working)):
            for t in titles:
                if working[i].startswith(t):
                    r = shlex.split(working[i])
                    if r[2:] == []:
                        rep.update({r[1]:r[0]})
                    else:
                        rep.update({r[1]:{r[0]:r[2:]}})
                    rem.append(working[i])
                    break
        
        for r in rem: # On retire les variables identifiées du bloc
            working.remove(r)
        
        if set([w[:4] for w in working if not w in end]) == set([' '*4]):
            working = [w.replace(' '*4,'',1) for w in working]
        
        if _end != None:
            end = _end
        
        working = '\n'.join(working)
        # On remet le bloc en forme de texte, pret pour un récurtion d'un bloc de plus
        # haut niveau
        if _old_index == index or title == '':
            # On sort de la boucle si le bloc n'avance plus, pour éviter la boucle
            # infini
            for key in res:
                # on retire des titres qui aurait pu être rajoutés
                for e in end+('',):
                    if e in res and res == {e:[]}:
                        res = e
                    elif e in res and res[e] in [{}]:
                        res.pop(e)
        elif working == '':
            # On ajout les variables avant de sortir
            res.update({title:rep})
        else:
            # On ajoute le bloc et on entre en récurtion avec la partie non traitée du bloc
            res.update({title:parse(working,end=end,res=rep)})
        # On prépare à passer au bloc suivant ou finir la boucle
        _old_index = index
        index = stop+1
    for key in res.keys():
        # On retire des titres qui aurait pu être rajoutés
        for e in end+('',):
            if e in res and res == {e:[]}:
                res = e
            elif e in res and res[e] in [{}]:
                res.pop(e)
    return res

def prepare(dico):
    """ Function that take Fortigate backup configuration file parsed 
    in a JSON/dict and return a JSON/dict structure where every key 
    of higer level is an option or a next level config
    
    Here is an example:
        >>> config = 'config system interface\\n'+\\
        ... '    edit "port 1"\\n'+\\
        ... '        set vdom "root"\\n'+\\
        ... '        set ip 192.168.1.254 255.255.255.0\\n'+\\
        ... '        set allowaccess ping\\n'+\\
        ... '        set type physical\\n'+\
        ... '        set alias "Internet access"\\n'+\\
        ... '        set snmp-index 1\\n'+\\
        ... '    next\\n'+\\
        ... 'end'
        >>> res = parse(config)
        >>> rem = prepare(res)
        >>> rem.keys()
        ['conf']
        >>> rem['conf'].keys()
        ['config system interface']
        >>> rem['conf']['config system interface'].keys()
        ['conf']
        >>> rem['conf']['config system interface']['conf'].keys()
        ['edit "port 1"']
        >>> rem['conf']['config system interface']['conf']['edit "port 1"'].keys()
        ['opt']
        >>> rem['conf']['config system interface']['conf']['edit "port 1"']['opt']
        ['alias', 'allowaccess', None, 'ip', 'type', 'snmp-index', 'vdom']
    
    return a dict or None
    """
    
    #
    # On verifie les entrées
    #
    
    if type(dico) is dict:
        keys = {'conf':{},'opt':{}}
        _keys = set()
        
        #
        # On va trier pour chacunes des valeurs du dictionnaire
        #
        
        for key in dico:
            _key = shlex.split(key)
            if _key[0] in ['config','edit']: 
                # Si c'est un niveau des configurations plus élevées
                keys['conf'].update({key:dico[key]})
            else:
                # Sinon c'est une donnée
                try:
                    value = dico[key]['set']
                except:
                    value = str(dico[key])
                if type(value) is list:
                    keys['opt'].update({key:'\n'.join(value)})
                else:
                    keys['opt'].update({key:str(value)})
                _keys.update([key])
        
        #
        # On ajoute un entrée qui contient le titre des données
        #
        keys['opt'].update({None:list(_keys)})
        
        #
        # Pour chaque niveau de configuration plus élevé
        # On déclence une récurtion
        #
        
        conf = {key:prepare(keys['conf'][key]) for key in keys['conf']}
        keys.update({'conf':conf})
        
        #
        # On retire les entrée si elles n'ont pas reçu de données
        #
        
        if keys['opt'][None] == []:
            _ = keys.pop('opt')
        if keys['conf'] == {}:
            _= keys.pop('conf')
        return keys
    else:
        return None

def build_xls(dico):
    """ Function that take Fortigate backup configuration file parsed 
    and prepared in a JSON/dict and return a tuple of two lists :
        - rep : All data at current level
        - conf : All data with a higher level
    
    Here is an example:
        >>> config = 'config system interface\\n'+\\
        ... '    edit "port 1"\\n'+\\
        ... '        set vdom "root"\\n'+\\
        ... '        set ip 192.168.1.254 255.255.255.0\\n'+\\
        ... '        set allowaccess ping\\n'+\\
        ... '        set type physical\\n'+\
        ... '        set alias "Internet access"\\n'+\\
        ... '        set snmp-index 1\\n'+\\
        ... '    next\\n'+\\
        ... 'end'
        >>> res = parse(config)
        >>> rem = prepare(res)
        >>> rep , conf = build_xls(rem)
        >>> rep
        []
        >>> for item in conf:
        ...     item
        ['Names']
        ['config system interface']
        [None, 'Names', 'ip', 'alias', 'allowaccess', 'type', 'snmp-index', 
        'vdom']
        [None, 'edit "port 1"', '192.168.1.254\n255.255.255.0', 
        'Internet access', 'ping', 'physical', '1', 'root']
    
    return a tuple of two lists
    """
    
    #
    # On verifie les entrées
    #
    
    if type(dico) is dict:
        if set([key in ['conf','opt'] for key in dico]) == set([True]):
            rep = []
            conf = []
            if dico.has_key('opt') and type(dico['opt']) is dict:
                
                #
                # On trie les données
                #
                
                rep += [dico['opt'][None]]# On donne le titre des futurs
                # colonnes
                index = len(rep)
                dico['opt'] = {key:dico['opt'][key] for key in dico['opt'] if key}
                rep += [[]]
                
                #
                # On rempli chaque colonne en s'assurant d'avoir une 
                # valeur pour chaque titre, sinon on remplie avec une
                # valeur nulle
                #
                
                for key in rep[index-1]:
                    if key in rep[index-1]:
                        rep[index] += [dico['opt'][key]]
                    else:
                        rep[index] += [None]
            
            if dico.has_key('conf') and type(dico['conf']) is dict:
                
                #
                # On trie les données des configurations plus élevées
                #
                
                __rep = {}
                __conf = {}
                for key in dico['conf']:
                    
                    #
                    # On entre en récurtion pour chaque configuration
                    # de plus haut niveau
                    #
                    
                    rep__, conf__ = build_xls(dico['conf'][key])
                    __rep.update({key:rep__})
                    __conf.update({key:conf__})
                
                #
                # On trie les données de ce niveau en regroupant
                # l'ensemble des configuration de même niveau
                # (uniquement le le niveau N+1)
                #
                
                titles = set()
                for _index in __rep:
                    if __rep[_index] in [None,[]]:
                        continue
                    titles.update(__rep[_index][0])
                
                titles = list(titles)
                conf += [['Names']+titles]
                if len(titles):
                    for key in dico['conf']:
                        index = len(conf)
                        conf += [[key]]
                        for _key in titles:
                            if __rep[key] in [None,[]]:
                                conf[index] += [None]
                            else:
                                if _key in __rep[key][0]:
                                    conf[index] += [__rep[key][1]\
                                        [__rep[key][0].index(_key)]]
                                else:
                                    conf[index] += [None]
                
                #
                # On regarde si on a un configuration de niveau N+2
                #
                
                if __conf == {key:[] for key in dico['conf']}:
                    # Ici toutes les configuration de niveau N+1 ne 
                    # contenait que des données
                    pass
                else:
                    for key in __conf:
                        # Pour chaque configuration de niveau N+2,
                        # on prend l'ensemble de ces données que l'on
                        # décale d'une colonne après avoir rajouté un
                        # titre
                        if __conf[key] in [None,[]]:
                            pass
                        else:
                            conf += [[key]]
                            for line in __conf[key]:
                                conf += [[None]+line]
            return rep, conf
        else:
            return None, None
    else:
        return None, None

def Parsed_to_xls(dico, key, name, save = False, profile = None):
    """ Function that take Fortigate backup configuration file parsed 
    and prepared in a JSON/dict and an Openpyxl Workbook
    
    Here is an example:
        >>> config = 'config system interface\\n'+\\
        ... '    edit "port 1"\\n'+\\
        ... '        set vdom "root"\\n'+\\
        ... '        set ip 192.168.1.254 255.255.255.0\\n'+\\
        ... '        set allowaccess ping\\n'+\\
        ... '        set type physical\\n'+\
        ... '        set alias "Internet access"\\n'+\\
        ... '        set snmp-index 1\\n'+\\
        ... '    next\\n'+\\
        ... 'end'
        >>> res = parse(config)
        >>> key = ['firewall','vpn','proxy','webfilter',
        'router','ips','system','application','antivirus']
        >>> name = 'test.xlsx'
        >>> wb = Parsed_to_xls(res, key, name, save = False, profile = None)
        >>> [(cell,wb['system interface'].__dict__['_cells'][cell].value) 
        ... for cell in wb['system interface'].__dict__['_cells']]
        [((1, 2), u'ip'), ((2, 7), u'root'), ((1, 3), u'alias'), 
        ((2, 2), u'192.168.1.254\n255.255.255.0'), 
        ((1, 4), u'allowaccess'), ((2, 4), u'ping'), ((1, 5), u'type'),
        ((2, 6), 1), ((1, 6), u'snmp-index'), ((2, 1), u'edit "port 1"'),
        ((1, 7), u'vdom'), ((2, 3), u'Internet access'), 
        ((2, 5), u'physical'), ((1, 1), u'Names')]
    
    return a Workbook openpyxl or None, if an input is not valid
    """
    
    #
    # On réduit le dictionnaire à la ou les clés prédéfinit
    #
    
    if type(dico) is dict:
        if type(key) in [str,unicode]:
            keys = sorted([_key for _key in dico if key in _key])
        elif type(key) is list:
            keys = sorted([_key for _key in dico if True in [__key in _key for __key in key]])
        else:
            pass
            #return None
        if type(profile) is dict:
            if set([type(profile[group]) for group in profile]) == set([dict]):
                _profile = {}
                for group in profile:
                    for category in profile[group]: 
                        _profile.update({category:
                            Webfilter_category(category, profile[group][category], 
                            group)})
            else:
                pass
                #return None
        elif profile is None:
            pass
        else:
            return None
    else:
        return None
        
    #
    # On prépare le Workbook
    # 
    
    wb = Workbook()
    accueil = wb.active
    accueil.title = "Accueil"
    for _key in keys: # Pour chacune des clés retenues
        __key = _key.split(' ',1)[1]
        wb.create_sheet(__key) # On crée une feuille de travail
        rem = prepare(dico[_key]) # On prepare le dictionnaire 
        if profile and __key == 'webfilter profile': # Le cas échéant on 
            # rajoute les profiles webfilter
            for conf in rem['conf']:
                filters = rem['conf'][conf]['conf']['config ftgd-wf']['conf']['config filters']['conf']
                for filter in filters:
                    if filters[filter]:
                        if filters[filter]['opt'].has_key('category'):
                            filters[filter]['opt'][None] += ['category name','category group']
                            __profile = _profile.get(filters[filter]['opt']['category'],None)
                            if hasattr(__profile,'name') and hasattr(__profile,'group'):
                                filters[filter]['opt'].update({'category name':__profile.name,
                                    'category group':__profile.group})
                            else:
                                filters[filter]['opt'].update({'category name':'unknow',
                                    'category group':'unknow'})
                        rem['conf'][conf]['conf']['config ftgd-wf']['conf']['config filters']['conf'][filter]['opt'] = \
                            filters[filter]['opt']
                    
                    
        rep, conf = build_xls(rem) # On recolte les données 
        # correctement formatées
        if rep == None:
            continue 
        _j = 0
        for j in range(len(rep)): # Pour l'ensemble des données
            _j = j # On se souvient du nombre de lignes écrites
            for i in range(len(rep[j])): # Pour chaque valeur
                value = rep[j][i]
                if value.isdigit(): # On la prépare sous forme de nombre...
                    wb[__key][get_column_letter(1+i)+str(j+1)].value = int(value)
                else: # ... ou de chaîne de caractère
                    wb[__key][get_column_letter(1+i)+str(j+1)].value = str(value)
                    if '\n' in str(value): # On prend en compte les données sur
                        # plusieurs lignes
                        wb[__key][get_column_letter(1+i)+str(j+1)].alignment = Alignment(wrapText=True)
        for j in range(len(conf)): # Pour l'ensemble des configurations de plus
            # haut niveau
            for i in range(len(conf[j])): # Pour chaque valeur
                value = conf[j][i]
                if value == None: # On s'assure qu'elle ne soit pas nulle
                    pass
                elif value.isdigit(): # On la prépare sous forme de nombre...
                    # On prend en compte les lignes déjà écritent
                    wb[__key][get_column_letter(1+i)+str(j+_j+1)].value = int(value)
                else: # ... ou de chaîne de caractère
                    # On prend en compte les lignes déjà écritent
                    wb[__key][get_column_letter(1+i)+str(j+_j+1)].value = str(value)
                    if '\n' in str(value): # On prend en compte les données sur
                        # plusieurs lignes et les lignes déjà écritent
                        wb[__key][get_column_letter(1+i)+str(j+_j+1)].alignment = Alignment(wrapText=True)
    
    #
    # On retire les feuilles vides
    #
    
    empty_worksheet = []
    for sheet in wb.sheetnames[1:]:
        if len(wb[sheet].__dict__['_cells']) < 2:
            empty_worksheet.append(sheet)
    
    #
    # On notifie sur la page d'accueil qu'elle ont tout de même
    # été traitée
    #
    
    accueil['J5'].value = 'Empty sheets'
    for sheet in empty_worksheet:
        wb.remove_sheet(wb[sheet])
        accueil['J'+str(6+empty_worksheet.index(sheet))].value = sheet
    
    #
    # On créer un lien vers l'ensemble des feuilles (non vide)
    #
    
    accueil['B5'].value = 'Links'
    for sheet in wb.sheetnames[1:]:
        accueil['B'+str(5+wb.sheetnames.index(sheet))].value = sheet
        accueil['B'+str(5+wb.sheetnames.index(sheet))].hyperlink = os.path.split(name)[1] +\
            '#'+(sheet,"'"+sheet+"'")[' ' in sheet]+'!A1'
    
    #
    # On ajuste la largeur des colonnes 
    #
    
    for sheet in wb.sheetnames: # Pour chaque feuille
        column_widths = []
        try:
            for row in wb[str(sheet)]: #Pour chacune ligne
                for i, cell in enumerate(row): # Pour chaque cellule
                    if cell.value != None: # si elle n'est pas vide
                        if len(column_widths) > i: # et que ce n'est pas une
                            # cellule de la première ligne
                            if type(cell.value) in [str,unicode]:
                                if '\n' in str(cell.value.encode('UTF-8')):
                                    # Si c'est une chaîne de caractère
                                    for value in cell.value.encode('UTF-8').split('\n'):
                                        # On verifie pour chaque ligne si le texte
                                        # est plus grand que la valeur maximale
                                        # déjà enregistrée
                                        if len(value) > column_widths[i]:
                                            column_widths[i] = len(value)+1
                                else:
                                    if len(cell.value) > column_widths[i]:
                                        column_widths[i] = len(cell.value)+1
                        else:
                            # Sinon, 
                            column_widths += [max(10,len(cell.value)+1)]
                    else:
                        if len(column_widths) < i:
                            # et que c'est une cellule de la première ligne
                            column_widths += [10]
            
            
            for i, column_width in enumerate(column_widths):
                # On fixe la largeur des colonne pour les valeurs données
                wb[str(sheet)].column_dimensions[get_column_letter(i+1)].width = column_width
        except TypeError, e:
            if e.message == "iter() returned non-iterator of type 'tuple'":
                pass
            else:
                raise TypeError, e
    
    if save: # On enregistre le fichier au besoin
        wb.save(name)
    return wb

def isfloat(value):
    try:
        float(value)
        return True
    except ValueError:
        return False

def webfilter_version(profiles, version):
    versions = [v for v in profiles if isfloat(v) and v[0].isdigit()]
    range_versions = [v for v in profiles if not v in versions]
    if version in versions:
        return profiles[version], version
    else:
        versions += list(set([v[1:] for v in profiles if isfloat(v) and not v[0].isdigit()]))
        _range_versions = sorted([float(_v) for _v in versions]+[0.0])
        for range_version in range_versions:
            _min, _max = [], []
            if range_version.startswith('+'):
                __version = float(range_version)
                for __v in _range_versions:
                    if __v > __version:
                        _max += [__v]
                _min, _max = __version,min(_max)
                if _min <= float(version) < _max:
                    return profiles[range_version], range_version
            elif range_version.startswith('-'):
                __version = -float(range_version)
                for __v in _range_versions[::-1]:
                    if __v < __version:
                        _min += [__v]
                _min, _max = max(_min),__version
                if _min <= float(version) < _max:
                    return profiles[range_version], range_version
            else:
                raise Exception, 'Invalid Format for version (%s) is unsupported ' % range_version[0]
    return None,None

def main():
    import re
    files = os.listdir(args.dir) # On liste l'ensemble des fichiers du répertoire choisit
    _files = []
    for file in args.files: # On filtre pour l'ensemble des arguments files fournit en CLI
        p = re.compile('('+file.replace('.','[.]').replace('*','.*?')+')$')
        _files += [p.findall(file)[0] for file in files if p.findall(file) != []]
    
    files = _files
    if args.webfilter: 
        # On effectue un recherche des webfilter profile 
        print 'Searching for profiles in',os.path.basename(args.profile),':',
        try: # On teste tout d'abord le fichier fournit en option
            fp = open(args.profile,'r')
            profiles = json.load(fp)
            fp.close()
            print 'done'
        except Exception, e: # En cas d'erreur
            print 'An error occur when profile has been load'
            _continue = None
            if args.profile == LOCAL_DIR+os.sep+'webfilter_profiles.json':
                # Si le fichier est celui de la librairie 
                print 'Default file can\'t be open !' 
            while not _continue in ['y','n']:
                # On demande à l'utilisateur s'il souhaite continuer sans l'ajout
                # des noms et groupes de webfilter
                _continue = raw_input('would you continue without webfilter ? (y/n) ')
            if _continue == 'y':
                # On désactive l'option
                args.webfilter = False
            elif args.profile == LOCAL_DIR+os.sep+'webfilter_profiles.json':
                # Dans le cas ou l'erreur a eu lieu sur le fichier de la librairie 
                print 'Program stop !'
                raise e
            else:
                _continue = None
                while not _continue in ['y','n']:
                    # On propose d'utiliser le fichier de la librairie
                    _continue = raw_input('would you continue with default profile ? (y/n) ')
                if _continue == 'y':
                    fp = open(LOCAL_DIR+os.sep+'webfilter_profiles.json','r')
                    profiles = json.load(fp)
                    fp.close()
                    print 'done'
                else:
                    print 'Program stop !'
                    raise e
        print
    
    # On cherche tous les fichiers qui vont correspondre a critère définit
    for file in files: # Pour chaque fichier
        print 'Processing :',file
        
        #
        # Lecture du fichier
        #
        
        print '\tReading :',
        f = open(args.dir+file,'r')
        config = f.read()
        f.close()
        print 'done'
        headers = []# Identification des headers
        print '\tHeader :',
        for conf in config.split('\n'):
            if conf.startswith('#'):
                headers.append(conf)
            else: 
                break
        
        index = sum([len(header) for header in headers])+len(headers)
        print 'done'
        
        #
        # Parsing de la conf
        #
        
        print '\tParsing :',
        res = parse(config[index:],res = {})
        print 'done'
        
        if args.XLSX:
            
            #
            # Creation des fichiers xlsx
            #
            
            #
            # Si on active le remplacement des numéro de webfilter
            #
            profile, version = None, None
            if args.webfilter:
                print '\tSearching for webfilter profiles :',
                try:
                    # On recherche la version des les headers du fichiers de conf
                    _version = [header for header in headers if header.startswith('#config-version=')][0].split('-')[2]
                    version = str(int(version.split('.')[0]))+'.'+str(int(_version.split('.')[1]))
                except:
                    pass
                
                if version:
                    # on recherche parmi les profiles identifiées, celui qui correspond
                    # à la version présente
                    profile, version = webfilter_version(profiles,version)
                    if profile:
                        print 'done (a profile has been found for %s)' % version
                    else:
                        print 'done (None profile found !)'
                else:
                    print 'Error during version analysis, webfilter profile aborted for this file !'
            print '\tFormat XLS :',
            wb = Parsed_to_xls(res, args.keys, args.dir+os.sep+file[::-1].split('.',1)[1][::-1]+'.xlsx', 
                save = True, profile = profile)
            print 'done'
        
        if args.JSON:
            
            #
            # Ecriture des fichiers JSON
            #
            
            res.update({'headers':headers})
            print '\tFormat JSON :',
            str_ = json.dumps(res,indent=4, sort_keys=True, ensure_ascii=False)
            with open(args.dir+os.sep+file[::-1].split('.',1)[1][::-1]+'.json', 'w') as outfile:
                outfile.write(str_)
            print 'done'
        
        print

if __name__ == '__main__':
    main()
